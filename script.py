#!/usr/bin/env python3
import requests
import json
import csv
import argparse
from openpyxl import Workbook
from collections import OrderedDict


client_id = "xxxxxxxxxx-xxxx-xxxx-xxx-xxxxxxxxxxxx"
client_secret = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"

def gettoken():
    url = "https://id.sophos.com/api/v2/oauth2/token"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    dati = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "token"
    }
    response = requests.post(url, data=dati, headers=headers)
    if response.status_code == 200:
        dati = response.json()
        access_token = dati.get("access_token")
        #print(f"Ecco il token :", access_token)
    else:
        print(f"Error: {response.status_code} - {response.text}")
    return access_token

access_token = gettoken()

def getenantid(access_token):
    url = "https://api.central.sophos.com/whoami/v1"
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        dati = response.json()
        tenantid = dati.get("id")
        #print("Ecco il tenant id:", tenantid)
    else:
        print(f"Error: {response.status_code} - {response.text}")
    return tenantid

tenantid = getenantid(access_token)

def listfw(access_token,tenantid):
    url = "https://api-eu02.central.sophos.com/firewall/v1/firewalls"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "X-Tenant-ID": tenantid
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        datijson = response.json()
        #print("Ecco i dati:", datijson)
    else:
        print(f"Error: {response.status_code} - {response.text}")
    return datijson

datijson = listfw(access_token,tenantid)


def extract_info(data):
    extracted_data = []
    for item in data.get('items', []):
        cluster_status = None
        if "cluster" in item and isinstance(item["cluster"], dict) and "status" in item["cluster"]:
            cluster_status = item["cluster"]["status"]
        extracted_data.append({
            "id": item.get("id", None),
            "serialNumber": item.get("serialNumber", None),
            "model": item.get("model", None),
            "hostname": item.get("hostname", None),
            "name": item.get("name", None),
            "externalIpv4Addresses": item.get("externalIpv4Addresses", []),
            "status": cluster_status
        })
    return extracted_data

def save_to_excel(extracted_data, filename="output.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items"

    headers = ["id", "serialNumber", "status", "model", "hostname", "name", "externalIpv4Addresses"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx).value = header

    row_idx = 2
    for item in extracted_data:
        sheet.cell(row=row_idx, column=1).value = item["id"]
        sheet.cell(row=row_idx, column=2).value = item.get("serialNumber")
        sheet.cell(row=row_idx, column=3).value = item.get("status")
        sheet.cell(row=row_idx, column=4).value = item.get("model")
        sheet.cell(row=row_idx, column=5).value = item.get("hostname")
        sheet.cell(row=row_idx, column=6).value = item.get("name")
        sheet.cell(row=row_idx, column=7).value = ", ".join(item.get("externalIpv4Addresses", []))
        row_idx += 1

    workbook.save(filename)

    
extracted_data = extract_info(datijson)
save_to_excel(extracted_data, filename="FWList.xlsx")

#formatted_output = json.dumps(extracted_info, indent=2)
#print(formatted_output)

